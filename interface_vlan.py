import openpyxl
from IPy import IP
IP_PLANNING_EXCEL = openpyxl.load_workbook('C:/Users/alawn/Desktop/宏业大厦IP规划&连接关系.xlsx')
IP_PLANNING_SHEET = IP_PLANNING_EXCEL['ip_planning_sheet']
MGT_IP_SHEET = IP_PLANNING_EXCEL['mgt_ip_sheet']
CONNECTION_SHEET = IP_PLANNING_EXCEL['connection_sheet']

IP_ROW = IP_PLANNING_SHEET.max_row
MGT_ROW = MGT_IP_SHEET.max_row
CONNECTION_ROW = CONNECTION_SHEET.max_row
lay2_vlan='''
vlan $vlan
name $name
'''

lay3_config_master = '''
interface $vlan
description $description
ip address $ipaddress $netmask
ip access $ACL in
no ip redirects
no ip unreachables
standby version 2
standby $vlan ip $hsrpip
standby $vlan priority 120
standby $vlan preemt
'''


lay3_config_backup = '''
interface $vlan
description $description
ip address $ipaddress $netmask
ip access $ACL in
no ip redirects
no ip unreachables
standby version 2
standby $vlan ip $hsrpip
standby $vlan preemt
'''

ospf_config_basic_info = '''
router ospf 100
route-id $route_id
area $area_id
no passive-interface $up_link_port
no passive-interface $vlan10
'''


def get_dict(excel,row):
    dict_list = []
    for r in range (2,row+1):
        floor = excel.cell(row=r,column=5).value
        network = excel.cell(row=r,column=2).value
        function = excel.cell(row=r,column=4).value
        int_network = IP(network.split('/')[0]).int()
        netmask = str(IP(network).strNormal(2)).split('/')[1]
        hsrpip = IP(int_network+1).strNormal(0)
        masterip = IP(int_network+2).strNormal(0)
        backupip = IP(int_network+3).strNormal(0)
        vlan = excel.cell(row=r,column=1).value
        acl = ''
        if vlan in range(11, 12):
            acl = 'AP'
        elif vlan in range (19,30):
            acl = 'OA'
        elif vlan == 44:
            acl = 'Video'
        elif vlan in range(599, 603):
            acl = 'OA-Device'
        elif vlan in range (29,40):
            acl = 'TY'
        elif vlan in range (99,110):
            acl = 'VOIP'
        description = excel.cell(row=r,column=3).value
        function = excel.cell(row=r,column=4).value
        if floor== None:
            pass
        else:
            dict_list.append(Ip_dict(vlan,description,hsrpip,masterip,backupip,netmask,acl,function,floor).__dict__)
    return dict_list

def get_key(excel,row):
    master_key = []
    for r in range (2,row+1):
        floor = excel.cell(row=r,column=5).value
        if floor == None:
            pass
        else:
            master_key.append(floor)
    return list(set(master_key))

for key in get_key(IP_PLANNING_SHEET,IP_ROW):
    print('\n\n'+'%s floor master doa config' %key)
    for sub_dict in get_dict(IP_PLANNING_SHEET,IP_ROW):
        if key  == sub_dict['floor']:
            # vlan = lay2_vlan.replace('$vlan',str(sub_dict['vlan'])).replace('$name',sub_dict['description'])
            # masterconfig  = lay3_config_master.replace('$vlan',str(sub_dict['vlan'])).replace('$ipaddress',sub_dict['masterip']).replace('$netmask',sub_dict['netmask']).replace('$ACL',sub_dict['acl']).replace('$description',sub_dict['description'])
            # backconfig = lay3_config_backup.replace('$vlan', str(sub_dict['vlan'])).replace('$ipaddress', sub_dict[
            #     'backupip']).replace('$netmask', sub_dict['netmask']).replace('$ACL', sub_dict['acl']).replace('$description',sub_dict['description'])
            # print(vlan,masterconfig,backconfig)
            ospf_config_network_info = ('network %s 0.0.0.0 $area_id' %sub_dict['masterip'])
            print(ospf_config_network_info)