import pymysql
import openpyxl



#
# def ip_planning(project):
#     db = pymysql.connect(host='119.91.102.106', user='root', password='uz8954UZN', database='building_information')
#
#     cursor1 = db.cursor(cursor=pymysql.cursors.DictCursor, )
#     ip_planning = "select * from ip_planning where project = '%s' ORDER BY floor*1" %project
#     cursor1.execute(ip_planning)
#     ip_data = cursor1.fetchall()
#     return ip_data
#
# def connection(project):
#     db = pymysql.connect(host='119.91.102.106', user='root', password='uz8954UZN', database='building_information')
#
#     cursor1 = db.cursor(cursor=pymysql.cursors.DictCursor, )
#     info = "select * from connection_relation where project = '%s'" % project
#     cursor1.execute(info)
#     connection = cursor1.fetchall()
#     return connection
#
# def deivce_ip(project):
#     db = pymysql.connect(host='119.91.102.106', user='root', password='uz8954UZN', database='building_information')
#
#     cursor1 = db.cursor(cursor=pymysql.cursors.DictCursor, )
#     info = "select * from Manage_IP_assignments where project = '%s'" % project
#     cursor1.execute(info)
#     device_ip = cursor1.fetchall()
#     return device_ip
#
# def write_excel(project):
#     workbook = openpyxl.Workbook()
#     ip_sheet = workbook.create_sheet('ip_planning')
#     ip_sheet.append(['network','status','domain','vlan','func','description','acl','project','building_name','floor','bdr','type_of_workplace'])
#     mgt_sheet = workbook.create_sheet('mgt_ip')
#     mgt_sheet.append([''])
#     connect = workbook.create_sheet('connect')
#     for info in ip_planning(project):


# import  mysql_table_query
# def basic_device_info_dict(project):
#     doa_config_info = []
#     manage_ip_list = mysql_table_query.deivce_ip(project)
#     network = mysql_table_query.ip_planning(project)
#     connect = mysql_table_query.connection(project)
#     interconnection_list = []
#     downlinkconnection_list = []
#     uplinkconnection_list = []
#     for c in connect:
#         connection = {'floor': None, 'A_device': None, 'A_port': None, 'Z_device': None, 'Z_port': None}
#         if '-DOA-' in c['A_device'] and '-DOA-' in c['Z_device'] :
#             connection['floor'] = c['Z_floor']
#             connection['A_device'] = c['A_device']
#             connection['A_port'] = c['A_port']
#             connection['Z_device'] = c['Z_device']
#             connection['Z_port'] = c['Z_port']
#             interconnection_list.append(connection)
#
#         if '-DOA-' in c['A_device'] and '-DOA-' not in c['Z_device'] :
#             connection['floor'] = c['Z_floor']
#             connection['A_device'] = c['A_device']
#             connection['A_port'] = c['A_port']
#             connection['Z_device'] = c['Z_device']
#             connection['Z_port'] = c['Z_port']
#             downlinkconnection_list.append(connection)
#
#         layer3connection = {'layer3connection':{'floor': None, 'A_device': None, 'A_port': None,'A_ipaddress':None,'Z_device': None, 'Z_port': None,'Z_ipaddress':None}}
#         if '-COA' in c['A_device'] and '-DOA-' in c['Z_device']:
#             layer3connection['layer3connection']['floor'] = c['Z_floor']
#             layer3connection['layer3connection']['A_device'] = c['A_device']
#             layer3connection['layer3connection']['A_port'] = c['A_port']
#             layer3connection['layer3connection']['A_ipaddress'] = c['A_ip']
#             layer3connection['layer3connection']['Z_device'] = c['Z_device']
#             layer3connection['layer3connection']['Z_port'] = c['Z_port']
#             layer3connection['layer3connection']['Z_ipaddress'] = c['Z_ip']
#             uplinkconnection_list.append(layer3connection)
#     for ip in manage_ip_list:
#         interconnect = {'interconnect':[]}
#         for n in interconnection_list:
#             if ip['device_name'] == n['A_device'] or ip['device_name'] == n['Z_device']:
#                 interconnect['interconnect'].append(n)
#             if ip['floor'] == n['floor'] and '-DOA' in ip['device_name']:
#                 ip.update(interconnect)
#
#
#         d_downlinkconnect = {'downlinkconnect': []}
#         e_downlinkconnect = {'downlinkconnect': []}
#         for m in downlinkconnection_list:
#             if '-DOA-' in ip['device_name'] and  '-D-' in ip['device_name'] and '-D-' in m['A_device'] and ip['floor'] == m['floor'] :
#                 d_downlinkconnect['downlinkconnect'].append(m)
#                 ip.update(d_downlinkconnect)
#             if '-DOA-' in ip['device_name'] and  '-E-' in ip['device_name'] and '-E-' in m['A_device'] and ip['floor'] == m['floor'] :
#                 e_downlinkconnect['downlinkconnect'].append(m)
#                 ip.update(e_downlinkconnect)
#
#         for u in uplinkconnection_list:
#             if ip['device_name'] == u['layer3connection']['Z_device']:
#                 ip.update(u)
#
#
#         networklist = []
#
#         for n in network:
#             interface_vlan = {'vlan': None, 'network': None,'desc':None,'acl':None}
#             if str(n['floor'])+str(n['bdr']) == str(ip['floor'])+str(ip['bdr']):
#                 interface_vlan['vlan'] = n['vlan']
#                 interface_vlan['network'] = n['network']
#                 interface_vlan['desc'] = n['description']
#                 interface_vlan['acl'] = n['acl']
#
#
#                 networklist.append(interface_vlan)
#         if '-DOA-' not in ip['device_name']:
#             pass
#         else:
#             ip.update({'network':networklist})
#             doa_config_info.append(ip)
#
#     return doa_config_info
#
# doa = basic_device_info_dict(project)
# for d in doa:
#     print(d)






# import mysql_table_query
# from math import ceil
# from math import floor
# project = input('项目名称: ')
# #
# network = input('IP地址：')
#
#
# endpoint = mysql_table_query.endpoint(project)
#
# def calc_num_oa(num_oa_point):
#     num_oa_network = None
#     num = num_oa_point / 240
#     if num < 1:
#         num_oa_network = 1
#     else:
#         numsplit = '{:.2}'.format(num_oa_point / 240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             num_oa_network = floor(num)
#         if int(decimals[1]) > 2:
#             num_oa_network = ceil(num)
#     return num_oa_network
#
# def calc_num_ty(num_oa_point):
#     num_oa_network = None
#     num = num_oa_point / 240 * 0.5
#     if num < 1:
#         num_oa_network = 1
#     else:
#         numsplit = '{:.2}'.format(num_oa_point / 240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             num_oa_network = floor(num)
#         if int(decimals[1]) > 2:
#             num_oa_network = ceil(num)
#     return num_oa_network
#
#
# def num_of_network(project):
#     num_network = []
#     for network in mysql_table_query.endpoint(project):
#         if network['convergence'] =='N':
#             pass
#         else:
#             floor_network_oa = (network['dpoint'] + network['epoint'])
#             floor_network_ty = floor_network_oa * 0.5
#             floor_network_voip = ceil(network['vpoint'] / 240)
#
#
#             floor_network_num_dict = {'floor':network['floor'],'bdr':network['bdr'],'mgt':0.25,'ap_mgt':0.25,'video':0.25,
#                                       'oa_device':0.25,'geli':0.25,'oa':calc_num_oa(floor_network_oa),'ty':calc_num_ty(floor_network_ty),
#                                       'voip':floor_network_voip}
#             num_network.append(floor_network_num_dict)
#     return num_network


import itertools
# from math import ceil
# from math import floor
# import ipaddress
# import mysql_table_query
#
#
# project = input('项目名称: ')
# # #
# network = input('IP地址：')
#
# def calc_num_oa(num_oa_point):
#     num_oa_network = None
#     num = num_oa_point / 240
#     if num < 1:
#         num_oa_network = 1
#     else:
#         numsplit = '{:.2}'.format(num_oa_point / 240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             num_oa_network = floor(num)
#         if int(decimals[1]) > 2:
#             num_oa_network = ceil(num)
#     return num_oa_network
#
#
# def calc_num_ty(num_oa_point):
#     num_oa_network = None
#     num = num_oa_point / 240 * 0.5
#     if num < 1:
#         num_oa_network = 1
#     else:
#         numsplit = '{:.2}'.format(num_oa_point / 240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             num_oa_network = floor(num)
#         if int(decimals[1]) > 2:
#             num_oa_network = ceil(num)
#     return num_oa_network
#
#
# def num_of_network(project):
#     num_network = []
#     for network in mysql_table_query.endpoint(project):
#         if network['convergence'] == 'N':
#             pass
#         else:
#             floor_network_oa = (network['dpoint'] + network['epoint'])
#             floor_network_ty = floor_network_oa * 0.5
#             floor_network_voip = ceil(network['vpoint'] / 240)
#
#             floor_network_num_dict = {'floor': network['floor'], 'bdr': network['bdr'], 'mgt': 0.25, 'ap_mgt': 0.25,
#                                       'video': 0.25,
#                                       'oa_device': 0.25, 'geli': 0.25, 'oa': calc_num_oa(floor_network_oa),
#                                       'ty': calc_num_ty(floor_network_ty),
#                                       'voip': floor_network_voip}
#             num_network.append(floor_network_num_dict)
#     return num_network
#
#
# def cacl_public(project):
#     public_num = 0
#     for public in num_of_network(project):
#         num = public['mgt'] + public['ap_mgt'] + public['video'] + public['oa_device'] + public['geli']
#         public_num += num
#     return ceil(public_num)
#
#
# def cacl_floor_bdr_num(project):
#     bdr_list = []
#     for mgt in mysql_table_query.endpoint(project):
#         bdr_list.append('-'.join((str(mgt['floor']), str(mgt['bdr']))))
#     return bdr_list
#
#
# def cacl_oa(project):
#     num_oa = 0
#     for oa in num_of_network(project):
#         num_oa += oa['oa']
#     return ceil(num_oa)
#
#
# def cacl_ty(project):
#     num_ty = 0
#     for ty in num_of_network(project):
#         num_ty += ty['ty']
#     return ceil(num_ty)
#
#
# def cacl_voip(project):
#     num_voip = 0
#     for voip in num_of_network(project):
#         num_voip += voip['voip']
#     return ceil(num_voip)
#
#
# def wire_network_num(project):
#     wire_network_num = cacl_public(project) + cacl_oa(project) + cacl_ty(project) + cacl_voip(project)
#     return wire_network_num
#
#
# def building_area(project):
#     area = 0.00
#     for n in mysql_table_query.endpoint(project):
#         area = area + n['area']
#     return area
#
#
# def calc_wifi_network_num(project):
#     area = building_area(project)
#     officewifi_network = ceil(area / 13 / 240) * 2
#     staffwifi_network = ceil(area / 13 / 240) * 3
#     guestwifi_network = ceil(area / 13 / 240 / 4)
#     labwifi_network = ceil(area / 13 / 240 / 25)
#     return {'office-wifi': officewifi_network, 'staff-wifi': staffwifi_network, 'guest-wifi': guestwifi_network,
#             'lab-wifi': labwifi_network, 'staffv6-only': 1, 'staffv6-daul': 1}
#
#
# def network_class(network, project):
#     pubilc_network_list = []
#     pubilc_subnetwork_list = []
#     core_ip_list = []
#     loopback = []
#     mgt_network = []
#     ip_address = ipaddress.ip_network(network).subnets(new_prefix=24)
#     connect_ip = (len(cacl_floor_bdr_num(project))) * 8 / 224
#     if connect_ip < 1:
#         core_network = ip_address.__next__()
#         mgt_network.append(core_network)
#         for ip in core_network.subnets(new_prefix=30):
#             core_ip_list.append(ip)
#     else:
#         numbers_of_connect_ip = ceil(connect_ip)
#         while numbers_of_connect_ip != 0:
#             core_network = ip_address.__next__()
#             mgt_network.append(core_network)
#             for ip in core_network.subnets(new_prefix=30):
#                 core_ip_list.append(ip)
#             numbers_of_connect_ip = numbers_of_connect_ip - 1
#     core_ip = itertools.product(core_ip_list)
#     loopbackip = 0
#     while loopbackip <= 15:
#         loopback.append(core_ip.__next__())
#         loopbackip = loopbackip + 1
#     network_class_dict = {'mgt': mgt_network, 'loopback': loopback, 'connection_ip': core_ip, 'public': None,
#                           'normal': None}
#     n = cacl_public(project)
#     while n != 0:
#         n = n - 1
#         pubilc_network_list.append(ip_address.__next__())
#
#     for pu in pubilc_network_list:
#         for ip in ipaddress.ip_network(pu).subnets(new_prefix=26):
#             pubilc_subnetwork_list.append(ip)
#     network_class_dict['public'] = (publicip for publicip in pubilc_subnetwork_list)
#     normal_network_list = []
#     for i in ip_address:
#         normal_network_list.append(i)
#     network_class_dict['normal'] = (ip for ip in normal_network_list)
#     return network_class_dict
#
#
# #
# def mgt_num(project):
#     public_dict_list = []
#     function_list = ['网络设备管理', 'AP网', '会议设备网', '行政设备网', '隔离VLAN']
#     description_list = ['MGT', 'AP-MGT', 'Video', 'OA-Device', 'GELI']
#     for fun, des in zip(function_list, description_list):
#         for floor_bdr in cacl_floor_bdr_num(project):
#             if des == 'MGT':
#                 vlan = 10
#             elif des == 'AP-MGT':
#                 vlan = 11
#             elif des == 'Video':
#                 vlan = 44
#             elif des == 'OA-Device':
#                 vlan = 600
#             elif des == 'GELI':
#                 vlan = 666
#             floor_bdr_split = floor_bdr.split('-')
#             public_dict = {'vlan': vlan, 'floor': str(floor_bdr_split[0]), 'bdr': str(floor_bdr_split[1]),
#                            'network': '', 'fun': fun, 'desc': des}
#             public_dict_list.append(public_dict)
#     public_dict_generation = (p for p in public_dict_list)
#     return public_dict_generation
#
#
# def generation_netwrok_dict(project):
#     return {'public': cacl_public(project), 'oa': cacl_oa(project), 'ty': cacl_ty(project), 'voip': cacl_voip(project)}
#
#
# def network_assign(project):
#     normal_network_list = []
#     for floor in num_of_network(project):
#         vlan_oa = 19
#         num = 0
#         while floor['oa'] != 0:
#             floor['oa'] = floor['oa'] - 1
#             vlan_oa = vlan_oa + 1
#             num = num + 1
#             normal_network_list.append(
#                 {'vlan': vlan_oa, 'floor': str(floor['floor']), 'bdr': str(floor['bdr']), 'network': None,
#                  'fun': '有线办公网', 'desc': ('OA-' + str(num))})
#
#     for floor in num_of_network(project):
#         vlan_oa = 29
#         num = 0
#         while floor['ty'] != 0:
#             floor['ty'] = floor['ty'] - 1
#             vlan_oa = vlan_oa + 1
#             num = num + 1
#             normal_network_list.append(
#                 {'vlan': vlan_oa, 'floor': str(floor['floor']), 'bdr': str(floor['bdr']), 'network': None,
#                  'fun': '有线体验网', 'desc': ('TY-' + str(num))})
#
#     for floor in num_of_network(project):
#         vlan_oa = 99
#         num = 0
#         while floor['voip'] != 0:
#             floor['voip'] = floor['voip'] - 1
#             vlan_oa = vlan_oa + 1
#             num = num + 1
#             normal_network_list.append(
#                 {'vlan': vlan_oa, 'floor': str(floor['floor']), 'bdr': str(floor['bdr']), 'network': None,
#                  'fun': 'VOIP网', 'desc': ('VOIP-' + str(num))})
#
#     normal_network_list.append({'vlan': 10, 'floor': None, 'bdr': None, 'network': None, 'fun': 'xzjk',
#                                 'desc': 'xzjk-mgt'})
#     normal_network_list.append({'vlan': 90, 'floor': None, 'bdr': None, 'network': None, 'fun': 'xzjk',
#                                 'desc': 'xzjk-mgt'})
#     wifi_start_vlan = 19
#     all_wifi_network_num = (
#             calc_wifi_network_num(project)['office-wifi'] + calc_wifi_network_num(project)['staff-wifi']
#             + calc_wifi_network_num(project)['guest-wifi'] + calc_wifi_network_num(project)['lab-wifi'] +
#             calc_wifi_network_num(project)['staffv6-only'] + calc_wifi_network_num(project)['staffv6-daul'])
#     office_range = (office for office in range(calc_wifi_network_num(project)['office-wifi']))
#     staff_range = (staff for staff in range(calc_wifi_network_num(project)['staff-wifi']))
#     guest_range = (guest for guest in range(calc_wifi_network_num(project)['guest-wifi']))
#     normal_network_list.append({'vlan': 10, 'floor': None, 'bdr': None, 'network': None, 'fun': '无线核心管理段',
#                                 'desc': 'mgt'})
#     while all_wifi_network_num != 0:
#         wifi_start_vlan = wifi_start_vlan + 1
#         all_wifi_network_num = all_wifi_network_num - 1
#         try:
#             office = {'vlan': wifi_start_vlan, 'floor': None, 'bdr': None, 'network': None, 'fun': 'office',
#                       'desc': ('Office-WiFi-' + str(int(office_range.__next__()) + 1))}
#             normal_network_list.append(office)
#         except StopIteration:
#             try:
#                 staff = {'vlan': wifi_start_vlan, 'floor': None, 'bdr': None, 'network': None, 'fun': 'staff',
#                          'desc': ('Staff-WiFi-' + str(int(staff_range.__next__()) + 1))}
#                 normal_network_list.append(staff)
#             except StopIteration:
#                 try:
#                     guest = {'vlan': wifi_start_vlan, 'floor': None, 'bdr': None, 'network': None, 'fun': 'guest',
#                              'desc': ('Guest-WiFi-' + str(int(guest_range.__next__()) + 1))}
#                     normal_network_list.append(guest)
#                 except StopIteration:
#                     try:
#                         normal_network_list.append(
#                             {'vlan': wifi_start_vlan, 'floor': None, 'bdr': None, 'network': None, 'fun': 'lab',
#                              'desc': 'Lab-WiFi-1'})
#                     except EOFError:
#                         pass
#     normal_network_list.append({'vlan': 900, 'floor': None, 'bdr': None, 'network': None, 'fun': 'staffv6only',
#                                 'desc': 'staffv6only'})
#     normal_network_list.append({'vlan': 901, 'floor': None, 'bdr': None, 'network': None, 'fun': 'staffv6daul',
#                                 'desc': 'staffv6daul'})
#
#     normal_network_info = (i for i in normal_network_list)
#     return normal_network_info
#
#
# def acl(ipinfo):
#     acl = ''
#     if ipinfo == 'AP网':
#         acl = 'AP'
#     elif ipinfo == '会议设备网':
#         acl = 'Video'
#     elif ipinfo == '行政设备网':
#         acl = 'OA-Device'
#     elif ipinfo == '隔离VLAN':
#         acl = 'GELI'
#     elif ipinfo == '有线办公网':
#         acl = 'OA'
#     elif ipinfo == '有线体验网':
#         acl = 'TY'
#     elif ipinfo == 'VOIP网':
#         acl = 'VOIP'
#     return acl
#
#
# #
#
#
# def all_network_num(project, network):
#     all_network_num = (wire_network_num(project) + calc_wifi_network_num(project)['office-wifi'] +
#                        calc_wifi_network_num(project)['staff-wifi']
#                        + calc_wifi_network_num(project)['guest-wifi'] + calc_wifi_network_num(project)['lab-wifi'] +
#                        calc_wifi_network_num(project)['staffv6-only'] + calc_wifi_network_num(project)['staffv6-daul'])
#     new_networks = ipaddress.ip_network(network).subnets(new_prefix=24)
#     assign = [ip for ip in new_networks]
#     if all_network_num > len(assign):
#         print('IP地址分配不足')
#     else:
#         print('IP地址分配充足,利用率' + '{:.2%}'.format(all_network_num / len(assign)))
#
#
# def generation_ip_planning(network, project):
#     ip_planning_list = []
#     # #IP规划
#     core_ipaddress = network_class(network, project)['mgt']
#     core_network = len(core_ipaddress)
#     core_bdr_floor = mysql_table_query.workplace_info(project).pop(0)['core_bdr_floor']
#     if core_network <= 1:
#         ip_planning_list.append(
#             {'network': [core_ipaddress], 'status': '启用', 'domain': None, 'vlan': None, 'func': ['核心网段'],
#              'description': ['interconnection'], 'acl': None, 'project': project, 'building_name': None,
#              'floor': [core_bdr_floor], 'bdr': [1]})
#     else:
#         for n in core_ipaddress:
#             core_ip_dict = {'network': [n], 'status': '启用', 'domain': None, 'vlan': None, 'func': ['核心网段'],
#                             'description': ['interconnection'], 'acl': None, 'project': project, 'building_name': None,
#                             'floor': [core_bdr_floor], 'bdr': [1]}
#             ip_planning_list.append(core_ip_dict)
#
#     public_dic = mgt_num(project)
#     public_network_list = network_class(network, project)['public']
#     for n in public_network_list:
#         try:
#             public_network_basic_info = public_dic.__next__()
#             public_ip = {'network': [ipaddress.ip_network(n)], 'status': '启用', 'domain': None,
#                          'vlan': [public_network_basic_info['vlan']], 'func': [public_network_basic_info['fun']],
#                          'description': [public_network_basic_info['desc']],
#                          'acl': acl(public_network_basic_info['fun']),
#                          'project': project, 'building_name': None, 'floor': [public_network_basic_info['floor']],
#                          'bdr': [public_network_basic_info['bdr']]}
#             ip_planning_list.append(public_ip)
#         except StopIteration:
#             public_ip = {'network': [ipaddress.ip_network(n)], 'status': '未启用', 'domain': None,
#                          'vlan': None, 'func': None, 'description': None, 'acl': None,
#                          'project': project, 'building_name': None, 'floor': None, 'bdr': None}
#             ip_planning_list.append(public_ip)
#     normal_dic = network_assign(project)
#     normal_network_list = network_class(network, project)['normal']
#     for norip in normal_network_list:
#         try:
#             normal_ip = None
#             normal_network_basic_info = normal_dic.__next__()
#             if normal_network_basic_info['fun'] != 'xzjk':
#                 normal_ip = {'network': [ipaddress.ip_network(norip)], 'status': '启用', 'domain': None,
#                              'vlan': [normal_network_basic_info['vlan']], 'func': [normal_network_basic_info['fun']],
#                              'description': [normal_network_basic_info['desc']],
#                              'acl': acl(normal_network_basic_info['fun']), 'project': project, 'building_name': None,
#                              'floor': [normal_network_basic_info['floor']],
#                              'bdr': [normal_network_basic_info['bdr']]}
#             else:
#
#                 xzjk_network = ipaddress.ip_network(norip).subnets(new_prefix=26)
#
#                 for n in xzjk_network:
#                     normal_ip = {'network': [ipaddress.ip_network(n)], 'status': '启用', 'domain': None,
#                                  'vlan': [normal_network_basic_info['vlan']],
#                                  'func': [normal_network_basic_info['fun']],
#                                  'description': [normal_network_basic_info['desc']],
#                                  'acl': acl(normal_network_basic_info['fun']), 'project': project,
#                                  'building_name': None,
#                                  'floor': [normal_network_basic_info['floor']],
#                                  'bdr': [normal_network_basic_info['bdr']]}
#             ip_planning_list.append(normal_ip)
#         except StopIteration:
#             pass
#             normal_ip = {'network': [ipaddress.ip_network(norip)], 'status': '未启用', 'domain': None,
#                          'vlan': None, 'func': None, 'description': None, 'acl': None,
#                          'project': project, 'building_name': None, 'floor': None, 'bdr': None}
#
#             ip_planning_list.append(normal_ip)
#
#     return ip_planning_list

# generation_ip_planning(network,project)
# for i in generation_ip_planning(network,project):
#     print(i)

# project = input('项目名称: ')
# import ipaddress
# import mysql_table_query
# print(mysql_table_query.workplace_info(project))

# import mysql_table_query
# import device_name_prefix
# import ipaddress
# import device_port
# from math import ceil
#
#
# # project = input('项目名称: ')
# #
# #计算每层楼接入设备数量
# def device_number(iot):
#     if iot > 48:
#         xl = ceil(iot/48)
#     else:
#         xl = 1
#     return {'num_xl':xl}
# #
# def device_number_dict(project):
#     device_number_dict_list = []
#     for entry in mysql_table_query.endpoint(project):
#         device_number_dict = {'floor':entry['floor'],'bdr':entry['bdr']}
#         device_number_dict.update(device_number(entry['iot']))
#         device_number_dict_list.append(device_number_dict)
#     return device_number_dict_list
#
# def get_equipment_type(project):
#     type_list = []
#     for entry in mysql_table_query.equipment_type(project):
#         if entry['supplier'] == 'cisco':
#             equipment_type_acronym = 'C'
#         elif entry['supplier'] == 'hillstone':
#             equipment_type_acronym = 'S'
#         elif entry['supplier'] == 'h3c':
#             equipment_type_acronym = 'H'
#         elif entry['supplier'] == 'aruba':
#             equipment_type_acronym = 'A'
#         type_function = equipment_type_acronym+entry['name']+'-'+str(entry['function']).upper()
#         type_dict = {entry['function']:type_function}
#         type_list.append(type_dict)
#     return type_list
# # # #
#
# #
# #     # return all_mgt_list
#
# # #
# def generation_device_info_dict(project):
#     device_info_list = []
#     for entry in device_number_dict(project):
#         for type in get_equipment_type(project):
#             entry.update(type)
#         device_info_list.append(entry)
#     return device_info_list
#
#
#
# # #
# #
# def generation_floor_device_name(project):
#     devicelist = []
#     ccs_mgt_ip = ipaddress.IPv4Network(mysql_table_query.ccs_ip(project)[0]['network'])
#     # xl_mgt_ip = (ip for ip in ccs_mgt_ip[4:])
#     for entry in generation_device_info_dict(project):
#         xl_name = ['-'.join((device_name_prefix.device_prefix(mysql_table_query.workplace_info(project)[0]['city'],mysql_table_query.workplace_info(project)[0]['building_name']),('BDR'+str(entry['floor'])+str(entry['bdr']).rjust(2,'0')),'K',entry['xl'],str(num).rjust(2,'0'))) for num in range (1,entry['num_xl']+1)]
#         floor_device_dict = {'floor':entry['floor'],'bdr':entry['bdr'],'xl':xl_name}
#         devicelist.append(floor_device_dict)
#     return devicelist
#
# # #
# def get_xl_type(project):
#     port_assign = {'name': None, 'port_assign': None}
#     for entry in mysql_table_query.equipment_type(project):
#         if entry['function'] == 'xoa':
#             access_type = {'supplier':entry['supplier'],'type':entry['equipment_type']}
#             if access_type['supplier'] == 'cisco':
#                 port_assign['name'] = 'C'+str(entry['name'])
#                 port_assign['port_assign'] = device_port.cisco(access_type['type'])
#             elif access_type['supplier'] == 'h3c':
#                 port_assign['name'] = 'H'+str(entry['name'])
#                 port_assign['port_assign'] = device_port.h3c(access_type['type'])
#     return port_assign
#
# def get_xl_info(project):
#     xl_list = []
#     ccs_mgt_ip = ipaddress.IPv4Network(mysql_table_query.ccs_ip(project)[0]['network'])
#     xl_mgt_ip = (ip for ip in list(ccs_mgt_ip)[4:-2])
#     for n in generation_floor_device_name(project):
#         access_dict = {'floor':n['floor'],'bdr':n['bdr'],'xl':[]}
#         for name in n['xl']:
#             xl = {'floor':n['floor'],'bdr':n['bdr'],'name':name,'ip':xl_mgt_ip.__next__(),'netmask':ccs_mgt_ip.prefixlen,'gateway':ccs_mgt_ip[1],'port_assign':get_xl_type(project)['port_assign']}
#             if access_dict['floor'] == xl['floor']:
#                 access_dict['xl'].append(xl)
#         xl_list.append(access_dict)
#     return xl_list
#
# print(get_xl_info(project))
# for i in get_access_info(project):
#     print(i)

# project = input('项目名称: ')
# #
# # network = input('IP地址：')
#
# import mysql_table_query
#
#
# def special_floor(project):
#     special_floor = mysql_table_query.special_floor(project)
#     normal_floor = mysql_table_query.normal_floor(project)
#     new_floor_info_dict = []
#     for nor in normal_floor:
#         for spe in special_floor:
#             if nor['floor']+nor['bdr'] == spe['uplink_floor']+spe['uplink_bdr']:
#                 nor['dpoint'] +=  spe['dpoint']
#                 nor['epoint'] += spe['epoint']
#                 nor['vpoint'] += spe['vpoint']
#                 nor['area'] += spe['area']
#         new_floor_info_dict.append(nor)
#     return new_floor_info_dict
#
#
#
# def oa_calc_network(dpoint,epoint):
#     oa_num = None
#     oa = (dpoint+epoint)/240
#     if oa < 1:
#         oa_num = 1
#     else:
#         numsplit = '{:.2}'.format(oa/240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             oa_num = floor(oa)
#         if int(decimals[1]) > 2:
#             oa_num = ceil(oa)
#     return oa_num
#
# def ty_calc_network(dpoint,epoint):
#     ty_num = None
#     ty = (dpoint+epoint)/240*0.5
#     if ty < 1:
#         ty_num = 1
#     else:
#         numsplit = '{:.2}'.format(ty/240*0.85)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             ty_num = floor(ty)
#         if int(decimals[1]) > 2:
#             ty_num = ceil(ty)
#     return ty_num
#
# def voip_calc_network(vpoint):
#     voip_num = None
#     voip = vpoint/240
#     if voip < 1:
#         voip_num = 1
#     else:
#         numsplit = '{:.2}'.format(voip/240)
#         decimals = numsplit.split('.')
#         if int(decimals[1]) < 2:
#             voip_num = floor(voip)
#         if int(decimals[1]) > 2:
#             voip_num = ceil(voip)
#     return voip_num
#
# def ap_calc_network(area):
#     ap_num = None
#     ap = ceil(area/62*0.85)
#     if ap < 60 :
#         ap_num = 0.25
#     elif 61<ap<124:
#         ap_num = 0.5
#     elif 125 <ap<240:
#         ap_num = 1
#     return ap_num
#
# def num_of_switch(dpoint,epoint,vpoint,area):
#     rack_d = ceil(dpoint/44)
#     rack_e = ceil(epoint/44)
#     rack_v = ceil(vpoint/44)
#     rack_k = ceil(area/62*0.85/48)
#     num_of_switch = rack_d+rack_e+rack_v+rack_k
#     return num_of_switch
#
#
# for i in special_floor(project):
#     print(('BDR'+str(i['floor']).rjust(2,'0')+str(i['bdr']).rjust(2,'0')),oa_calc_network(i['dpoint'],i['epoint']),ty_calc_network(i['dpoint'],i['epoint']),voip_calc_network(i['vpoint']),ap_calc_network(i['area']))
#
#     print(num_of_switch(i['dpoint'],i['epoint'],i['vpoint'],i['area']))
import ipaddress
print(ipaddress.IPv4Interface('30.20.0.65/30').ip)
