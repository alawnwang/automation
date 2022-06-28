import os
import sys
import openpyxl
import mysql_table_query
from tkinter import filedialog

# project = '深圳光启未来'
#
# endpoint_info = mysql_table_query.endpoint(project)

device_info_file = [('excel file', '.xlsx'), ('excel file', '.xls')]

file = filedialog.askopenfilename(initialdir=os.getcwd(), title='Please select a files:', filetypes=device_info_file)
if not file:
    print('用户取消输入')
    sys.exit()
else:
    workbook = openpyxl.load_workbook(file)
    num_list = []
    for i in range(1, len(workbook.sheetnames) + 1):
        num_list.append(i)
    for n, s in zip(num_list, workbook.sheetnames):
        print('[' + str(n) + '] ' + s)


    def check_excle():
        dictlist = dict(zip(num_list, workbook.sheetnames))
        return dictlist


    check_excle()

    Choose_Sheet = input('choose:')


def choose_sheet():
    sheet = None
    for i in check_excle().keys():
        if Choose_Sheet == str(i):
            sheet = check_excle()[i]
    return sheet

choose_sheet()